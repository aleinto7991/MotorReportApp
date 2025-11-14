import os
import glob
from openpyxl import load_workbook

temp_dir = os.environ.get('TEMP') or os.environ.get('TMP') or r'C:\Windows\Temp'
pattern = os.path.join(temp_dir, 'Motor_Performance_Report_*.xlsx')
files = glob.glob(pattern)
if not files:
    print('No matching report files found in', temp_dir)
    raise SystemExit(1)

latest = max(files, key=os.path.getmtime)
print('Latest report:', latest)
wb = load_workbook(latest, read_only=True)
print('Sheets:', wb.sheetnames)

# Look for comparison-like sheets
candidates = [name for name in wb.sheetnames if 'comparison' in name.lower() or 'sap comparison' in name.lower()]
if not candidates:
    print('No comparison-named sheets found. Showing all sheets first rows instead.')
    candidates = wb.sheetnames[:3]

for sheet_name in candidates:
    print('\n--- Sheet:', sheet_name, '---')
    ws = wb[sheet_name]
    # Print headers (first row) and next 10 rows
    max_row = min(20, ws.max_row if ws.max_row else 20)
    for r in range(1, max_row+1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(30, ws.max_column)+1)]
        print(r, row_vals)

print('\nDone')
