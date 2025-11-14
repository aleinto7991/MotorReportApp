from __future__ import annotations

import logging
from pathlib import Path
from typing import cast
from zipfile import ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.config.app_config import AppConfig
from src.core.motor_report_engine import MotorReportApp
from src.services.directory_locator import DirectoryLocator


class _StubDirectoryLocator(DirectoryLocator):
    """Directory locator that leaves preconfigured paths untouched."""

    def __init__(self) -> None:
        super().__init__(logger=logging.getLogger(__class__.__name__))

    def apply_defaults(self, app_config: AppConfig) -> AppConfig:  # pragma: no cover - trivial passthrough
        # Ensure callers get exactly what the test configured while keeping parent dirs ready.
        if app_config.output_path:
            Path(app_config.output_path).parent.mkdir(parents=True, exist_ok=True)
        if app_config.log_path:
            Path(app_config.log_path).parent.mkdir(parents=True, exist_ok=True)
        return app_config


@pytest.mark.integration
def test_run_generates_excel_with_selected_sap(tmp_path: Path) -> None:
    """Full workflow: run() consumes INF/CSV data and produces an Excel report."""

    tests_dir = tmp_path / "performance"
    tests_dir.mkdir()

    # Prepare minimal INF/CSV pair for test number 12345 belonging to SAP123.
    inf_content = (
        "[General]\n"
        "Tipo motore = SAP123\n"
        "Data = 2024-01-02\n"
        "Tensione = 220\n"
        "Frequenza = 50Hz\n"
        "[Info_Aggiuntive]\n"
        "Note = Integration coverage\n"
    )
    (tests_dir / "12345.inf").write_text(inf_content, encoding="latin1")

    csv_content = (
        "Air Flow l/sec.;Vacuum Corrected mmH2O;Efficiency (%);"
        "Power Corrected Watts in;Speed RPM\n"
        "10;100;70;500;3000\n"
        "20;150;75;550;3100\n"
    )
    (tests_dir / "12345.csv").write_text(csv_content, encoding="latin1")

    output_path = tmp_path / "output" / "report.xlsx"
    log_path = tmp_path / "logs" / "run.log"

    config = AppConfig(
        tests_folder=str(tests_dir),
        output_path=str(output_path),
        log_path=str(log_path),
        include_noise=False,
        include_comparison=False,
        open_after_creation=False,
        sap_codes=["SAP123"],
    )

    app = MotorReportApp(config, directory_locator=_StubDirectoryLocator())

    app.run()

    assert output_path.exists(), "Excel report should be created"
    assert output_path.stat().st_size > 0, "Report file must not be empty"

    # Ensure MotorReportApp captured processed data for downstream consumers.
    assert len(app.all_motor_test_data) == 1
    test_data = app.all_motor_test_data[0]
    assert test_data.test_number == "12345"
    assert test_data.csv_data is not None and not test_data.csv_data.empty

    # Validate workbook contains our SAP identifier somewhere in the shared strings.
    with ZipFile(output_path) as workbook_zip:
        shared_strings = workbook_zip.read("xl/sharedStrings.xml").decode("utf-8")
    assert "SAP123" in shared_strings


@pytest.mark.integration
def test_run_includes_carichi_nominali_status_notes(tmp_path: Path) -> None:
    tests_dir = tmp_path / "performance"
    tests_dir.mkdir()

    test_lab_root = tmp_path / "test_lab"
    test_lab_root.mkdir()

    def write_inf_file(path: Path, sap_code: str, note: str = "") -> None:
        content = (
            "[General]\n"
            f"Tipo motore = {sap_code}\n"
            "Data = 2024-01-01\n"
            "Tensione = 230\n"
            "Frequenza = 50Hz\n"
            "[Info_Aggiuntive]\n"
            f"Note = {note}\n"
        )
        path.write_text(content, encoding="latin1")

    def write_csv_file(path: Path) -> None:
        csv_content = (
            "Air Flow l/sec.;Vacuum Corrected mmH2O;Efficiency (%);"
            "Power Corrected Watts in;Speed RPM\n"
            "10;100;70;500;3000\n"
            "20;150;75;550;3100\n"
        )
        path.write_text(csv_content, encoding="latin1")

    def create_test_lab_workbook(root: Path, stem: str) -> None:
        workbook = Workbook()
        scheda = cast(Worksheet, workbook.active)
        scheda.title = "Scheda SR"

        headers = ["", "Orifice", "Watt", "Air Watt"]
        for col_idx, value in enumerate(headers, start=1):
            scheda.cell(row=1, column=col_idx, value=value)

        data_rows = [
            ("Media", 12.0, 345.0, 56.0),
            ("Min", 11.0, 330.0, 50.0),
            ("Max", 13.0, 360.0, 59.0),
        ]
        for row_offset, row_values in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_values, start=1):
                scheda.cell(row=row_offset, column=col_idx, value=value)

        collaudo = cast(Worksheet, workbook.create_sheet("Collaudo SR"))
        collaudo["A5"] = "Media"
        for column_index in range(2, 14):
            collaudo.cell(row=5, column=column_index, value=float(column_index))

        workbook.save(root / f"{stem}.xlsx")

    sap_codes = ["6110820119", "6210820179"]
    for sap_code in sap_codes:
        base = sap_code
        variant = f"{sap_code}A"
        write_inf_file(tests_dir / f"{base}.inf", sap_code, note="Base test")
        write_inf_file(tests_dir / f"{variant}.inf", sap_code, note="Variant uses base data")
        write_csv_file(tests_dir / f"{base}.csv")
        create_test_lab_workbook(test_lab_root, variant)

    output_path = tmp_path / "output" / "report.xlsx"
    log_path = tmp_path / "logs" / "run.log"

    config = AppConfig(
        tests_folder=str(tests_dir),
        test_lab_root=str(test_lab_root),
        output_path=str(output_path),
        log_path=str(log_path),
        include_noise=False,
        include_comparison=False,
        open_after_creation=False,
        sap_codes=sap_codes,
    )

    app = MotorReportApp(config, directory_locator=_StubDirectoryLocator())

    app.run()

    status_by_test = {data.test_number: data.status_message for data in app.all_motor_test_data}
    assert status_by_test["6110820119A"] == "CSV data reused from 6110820119"
    assert status_by_test["6210820179A"] == "CSV data reused from 6210820179"

    tests_by_sap = {}
    carichi_workbooks = {}
    scheda_media_watt = {}

    for data in app.all_motor_test_data:
        tests_by_sap.setdefault(data.sap_code, []).append(data.test_number)
        if data.test_lab_summary and data.test_lab_summary.source_path:
            carichi_workbooks[data.test_number] = Path(data.test_lab_summary.source_path).name
            if data.test_lab_summary.scheda:
                media_row = data.test_lab_summary.scheda.rows.get("Media", {})
                scheda_media_watt[data.test_number] = media_row.get("Watt")

    assert set(tests_by_sap["6110820119"]) == {"6110820119", "6110820119A"}
    assert set(tests_by_sap["6210820179"]) == {"6210820179", "6210820179A"}

    assert carichi_workbooks["6110820119"] == "6110820119A.xlsx"
    assert carichi_workbooks["6210820179"] == "6210820179A.xlsx"

    assert scheda_media_watt["6110820119"] == pytest.approx(345.0)
    assert scheda_media_watt["6210820179"] == pytest.approx(345.0)

    with ZipFile(output_path) as workbook_zip:
        shared_strings = workbook_zip.read("xl/sharedStrings.xml").decode("utf-8")

    assert "Carichi nominali (Test-Lab" in shared_strings
    assert "CSV data reused from 6110820119" in shared_strings
    assert "CSV data reused from 6210820179" in shared_strings
    assert "Scheda SR Summary" in shared_strings
