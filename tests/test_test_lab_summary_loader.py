from __future__ import annotations

from pathlib import Path
from typing import cast

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.services.test_lab_summary_loader import TestLabSummaryLoader


def _create_test_lab_workbook(path: Path) -> None:
    """Create a minimal workbook containing the sections expected by the loader."""
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    scheda = cast(Worksheet, workbook.active)
    scheda.title = "Scheda SR"

    # Header row recognised by the extractor
    scheda["A5"] = "Label"
    scheda["B5"] = "Orifice"
    scheda["C5"] = "Watt"

    scheda["A6"] = "Media"
    scheda["B6"] = 12.3
    scheda["C6"] = 345.6

    collaudo = cast(Worksheet, workbook.create_sheet("Collaudo SR"))
    collaudo["A5"] = "Media"
    for column_index in range(2, 14):
        collaudo.cell(row=5, column=column_index, value=float(column_index))

    workbook.save(path)


def test_load_summary_prefers_exact_match_when_available(tmp_path: Path) -> None:
    base_dir = tmp_path / "CARICHI"
    year_dir = base_dir / "2024"

    exact_path = year_dir / "12345.xlsx"
    alias_path = year_dir / "12345A.xlsx"

    _create_test_lab_workbook(exact_path)
    _create_test_lab_workbook(alias_path)

    loader = TestLabSummaryLoader(str(base_dir))

    summary = loader.load_summary("12345")
    assert summary is not None
    assert summary.source_path is not None
    assert Path(summary.source_path).name == "12345.xlsx"
    assert summary.match_strategy == "exact"
    assert summary.matched_test_number == "12345"

    alias_summary = loader.load_summary("12345A")
    assert alias_summary is not None
    assert alias_summary.source_path is not None
    assert Path(alias_summary.source_path).name == "12345A.xlsx"
    assert alias_summary.match_strategy == "exact"


def test_load_summary_strict_matching(tmp_path: Path) -> None:
    """Verify that strict matching is enforced (no fallback to 'A' alias)."""
    base_dir = tmp_path / "CARICHI"
    year_dir = base_dir / "2024"

    alias_path = year_dir / "67890A.xlsx"
    _create_test_lab_workbook(alias_path)

    loader = TestLabSummaryLoader(str(base_dir))

    # Should NOT match 67890A when 67890 is requested
    summary = loader.load_summary("67890")
    assert summary is None

    # Should match 67890A when 67890A is requested
    alias_summary = loader.load_summary("67890A")
    assert alias_summary is not None
    assert alias_summary.source_path is not None
    assert Path(alias_summary.source_path).name == "67890A.xlsx"
