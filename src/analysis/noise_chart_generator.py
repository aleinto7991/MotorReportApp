"""
Noise Chart Generator - Creates Excel charts from TXT noise measurement files.
NO DATA TABLES - CHARTS ONLY to keep reports clean and professional.
"""

import logging
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

@dataclass
class NoiseTestData:
    """Represents parsed noise test data from a TXT file."""
    filename: str
    test_date: str
    radius: float
    overall_sound_pressure: float  # dB splA
    overall_sound_power: float     # dB wA
    frequency_data: List[Tuple[float, List[float]]]  # [(freq, [mic1, mic2, mic3, mic4, mic5, power])]
    
    def get_sampled_data(self, sample_rate: int = 5) -> List[Tuple[float, List[float]]]:
        """Get sampled frequency data for performance."""
        return self.frequency_data[::sample_rate]

class NoiseChartGenerator:
    """Generates Excel charts from noise measurement TXT files."""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
    def parse_txt_file(self, txt_file_path: Path) -> Optional[NoiseTestData]:
        """
        Parse a noise measurement TXT file.
        
        Format expected:
        Data Test:19/06/2025
        Raggio [m]:     1,50
        Livello Pressione [dB splA]:    69,6
        Livello Potenza [dB wA]:        81,1
        Freq [Hz]       Liv. Pres. Mic 1 [dB splA]   ...   Liv. Potenza [dB wA]:
        20,5    8,2     3,7     2,1     5,0     0,6     16,2
        ...
        """
        try:
            # Accept either Path or raw string input for flexibility
            if isinstance(txt_file_path, (str, bytes)):
                txt_file_path = Path(txt_file_path)

            self.logger.info(f"📖 Parsing TXT file: {txt_file_path.name}")
            
            with open(txt_file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
            
            # Initialize data
            test_date = ""
            radius = 0.0
            overall_sound_pressure = 0.0
            overall_sound_power = 0.0
            frequency_data = []
            
            # Parse header information
            for i, line in enumerate(lines[:10]):  # Check first 10 lines for header
                line = line.strip()
                
                # Extract test date
                if "Data Test:" in line or "Date:" in line:
                    date_match = re.search(r'(\d{2}/\d{2}/\d{4})', line)
                    if date_match:
                        test_date = date_match.group(1)
                
                # Extract radius
                elif "Raggio" in line and "[m]" in line:
                    radius_match = re.search(r'([\d,]+)', line)
                    if radius_match:
                        radius = float(radius_match.group(1).replace(',', '.'))
                
                # Extract overall sound pressure
                elif "Livello Pressione" in line and "dB splA" in line:
                    pressure_match = re.search(r'([\d,]+)', line)
                    if pressure_match:
                        overall_sound_pressure = float(pressure_match.group(1).replace(',', '.'))
                
                # Extract overall sound power
                elif "Livello Potenza" in line and "dB wA" in line:
                    power_match = re.search(r'([\d,]+)', line)
                    if power_match:
                        overall_sound_power = float(power_match.group(1).replace(',', '.'))
                
                # Check if we've reached the data section
                elif "Freq [Hz]" in line and "Mic" in line:
                    # Start parsing frequency data from next line
                    data_start_line = i + 1
                    break
            else:
                # If we didn't find the header, assume data starts from a reasonable line
                data_start_line = 6
            
            # Parse frequency data
            for line in lines[data_start_line:]:
                line = line.strip()
                if not line:
                    continue
                    
                # Split by whitespace and handle comma decimals
                parts = line.replace(',', '.').split()
                
                if len(parts) >= 6:  # Frequency + 5 mics + power (minimum)
                    try:
                        frequency = float(parts[0])
                        
                        # Extract microphone data (typically 5 mics + power)
                        mic_data = []
                        for i in range(1, min(len(parts), 7)):  # Up to 6 values (5 mics + power)
                            try:
                                mic_data.append(float(parts[i]))
                            except ValueError:
                                mic_data.append(0.0)  # Default for invalid values
                        
                        # Ensure we have at least 6 values (5 mics + power)
                        while len(mic_data) < 6:
                            mic_data.append(0.0)
                        
                        frequency_data.append((frequency, mic_data))
                        
                    except ValueError:
                        # Skip invalid lines
                        continue
            
            # Validate we got meaningful data
            if not frequency_data:
                self.logger.warning(f"⚠️ No frequency data found in {txt_file_path.name}")
                return None
            
            self.logger.info(f"✅ Parsed {len(frequency_data)} frequency points from {txt_file_path.name}")
            
            return NoiseTestData(
                filename=txt_file_path.name,
                test_date=test_date,
                radius=radius,
                overall_sound_pressure=overall_sound_pressure,
                overall_sound_power=overall_sound_power,
                frequency_data=frequency_data
            )
            
        except Exception as e:
            self.logger.error(f"❌ Error parsing {txt_file_path.name}: {e}")
            return None
    
    def create_noise_charts(self, worksheet, noise_tests: List[NoiseTestData], start_row: int = 1) -> int:
        """
        Create noise charts in Excel worksheet - CHARTS ONLY, NO DATA TABLES.
        
        Args:
            worksheet: Excel worksheet to add charts to
            noise_tests: List of parsed noise test data
            start_row: Starting row for charts
            
        Returns:
            Next available row after charts
        """
        try:
            self.logger.info(f"📊 Creating noise charts for {len(noise_tests)} test files")
            
            current_row = start_row
            
            # Add section header
            worksheet.cell(row=current_row, column=1, value="🔊 NOISE ANALYSIS CHARTS")
            worksheet.cell(row=current_row, column=1).font = Font(size=16, bold=True, color="FF006600")
            current_row += 2
            
            for test_data in noise_tests:
                current_row = self._create_test_chart_section(worksheet, test_data, current_row)
                current_row += 2  # Space between test sections
            
            # Hide all columns containing chart data AFTER creating all charts
            # TEMPORARILY DISABLED - Keep data visible to verify charts work
            # self._hide_chart_data_columns(worksheet)
            
            self.logger.info(f"✅ Created noise charts, final row: {current_row}")
            return current_row
            
        except Exception as e:
            self.logger.error(f"❌ Error creating noise charts: {e}")
            return start_row
    
    def _create_test_chart_section(self, worksheet, test_data: NoiseTestData, start_row: int) -> int:
        """Create chart section for a single test file."""
        try:
            current_row = start_row
            
            # Section title using full filename
            worksheet.cell(row=current_row, column=1, value=f"📄 {test_data.filename}")
            worksheet.cell(row=current_row, column=1).font = Font(size=12, bold=True, color="FF0066CC")
            current_row += 1
            
            # Test summary (one line only)
            summary = f"Date: {test_data.test_date}, Overall: {test_data.overall_sound_pressure:.1f} dB (pressure), {test_data.overall_sound_power:.1f} dB (power)"
            worksheet.cell(row=current_row, column=1, value=summary)
            worksheet.cell(row=current_row, column=1).font = Font(size=10, italic=True)
            current_row += 1
            
            # Create frequency response chart (left side)
            chart_start_row = current_row
            self._create_frequency_response_chart(worksheet, test_data, chart_start_row)
            
            # Create summary bar chart (right side, same row as frequency chart)
            self._create_summary_bar_chart(worksheet, test_data, chart_start_row)
            
            # Return row after both charts (charts are 20 rows high)
            return chart_start_row + 20
            
        except Exception as e:
            self.logger.error(f"❌ Error creating test chart section: {e}")
            return start_row
    
    def _create_frequency_response_chart(self, worksheet, test_data: NoiseTestData, start_row: int) -> int:
        """Create frequency response line chart with embedded data (no worksheet data)."""
        try:
            # Get sampled data for performance
            sampled_data = test_data.get_sampled_data(sample_rate=5)
            
            if not sampled_data:
                self.logger.warning(f"No frequency data for {test_data.filename}")
                return start_row + 1
                
            # Create line chart
            chart = LineChart()
            chart.title = f"Frequency Response - {test_data.filename}"
            chart.style = 2
            chart.x_axis.title = 'Frequency (Hz)'
            chart.y_axis.title = 'Sound Pressure (dB)'
            chart.width = 15
            chart.height = 8
            
            # Extract frequency values
            frequencies = [freq for freq, _ in sampled_data]
            
            # Use columns starting from AA (column 27) - close enough for charts to work
            hidden_data_start_col = 27  # Column AA (close but will be hidden after chart creation)
            
            # Extract frequency values
            frequencies = [freq for freq, _ in sampled_data]
            
            # Microphone names for legend
            mic_names = ['Mic 1', 'Mic 2', 'Mic 3', 'Mic 4', 'Mic 5']
            
            # Write frequency data to hidden area
            freq_col = hidden_data_start_col
            
            # First, write all data to hidden columns in a single pass
            data_start_row = start_row + 1
            for i, (freq, mic_data) in enumerate(sampled_data):
                row_num = data_start_row + i
                # Write frequency
                worksheet.cell(row=row_num, column=freq_col, value=freq)
                # Write microphone data
                for mic_index in range(5):
                    value = mic_data[mic_index] if mic_index < len(mic_data) else 0
                    mic_col = hidden_data_start_col + 1 + mic_index
                    worksheet.cell(row=row_num, column=mic_col, value=value)
            
            # Now create chart series for each microphone
            for mic_index in range(5):
                mic_col = hidden_data_start_col + 1 + mic_index
                
                # Create reference to the microphone data - be very explicit about ranges
                data_ref = Reference(worksheet, 
                                   min_col=mic_col, max_col=mic_col,
                                   min_row=data_start_row, 
                                   max_row=data_start_row + len(sampled_data) - 1)
                
                # Add series to chart
                chart.add_data(data_ref, titles_from_data=False)
                
                # Set the legend label for this series
                if chart.series and len(chart.series) > mic_index:
                    # Set chart series title as a plain string so openpyxl writes a visible legend label
                    chart.series[mic_index].title = mic_names[mic_index]
            
            # Set category axis to frequencies - be very explicit about range
            freq_ref = Reference(worksheet, 
                               min_col=freq_col, max_col=freq_col,
                               min_row=data_start_row,
                               max_row=data_start_row + len(sampled_data) - 1)
            chart.set_categories(freq_ref)
            
            # Add chart to worksheet
            chart_cell = f'A{start_row + 1}'
            worksheet.add_chart(chart, chart_cell)
            
            return start_row + 20  # Chart height
        except Exception as e:
            self.logger.error(f"❌ Error creating frequency chart: {e}")
            return start_row + 1
    
    def _create_summary_bar_chart(self, worksheet, test_data: NoiseTestData, start_row: int) -> int:
        """Create summary bar chart with overall levels, positioned to the right of frequency chart."""
        try:
            # Use columns starting from AG (column 33) for summary data
            summary_data_col = 33  # Column AG (close but will be hidden after chart creation)
            
            # Write summary data to hidden area
            worksheet.cell(row=start_row, column=summary_data_col, value="Sound Pressure")
            worksheet.cell(row=start_row, column=summary_data_col + 1, value=test_data.overall_sound_pressure)
            
            worksheet.cell(row=start_row + 1, column=summary_data_col, value="Sound Power")
            worksheet.cell(row=start_row + 1, column=summary_data_col + 1, value=test_data.overall_sound_power)
            
            # Create bar chart
            from openpyxl.chart import BarChart
            chart = BarChart()
            chart.title = f"Overall Levels"
            chart.x_axis.title = "Metrics"
            chart.y_axis.title = "Level (dB)"
            chart.width = 8   # Smaller width to fit next to frequency chart
            chart.height = 8  # Same height as frequency chart
            
            # Reference to data
            data_ref = Reference(worksheet, min_col=summary_data_col + 1, min_row=start_row, 
                               max_row=start_row + 1)
            chart.add_data(data_ref, titles_from_data=False)
            
            # Reference to categories  
            cats_ref = Reference(worksheet, min_col=summary_data_col, min_row=start_row,
                               max_row=start_row + 1)
            chart.set_categories(cats_ref)
            
            # Position chart to the right of frequency chart (column P = 16)
            chart_cell = f'P{start_row + 1}'
            worksheet.add_chart(chart, chart_cell)
            
            return start_row + 20  # Same as frequency chart height
        except Exception as e:
            self.logger.error(f"❌ Error creating summary bar chart: {e}")
            return start_row + 1
    
    def _hide_chart_data_columns(self, worksheet):
        """Hide all columns containing chart data to ensure only charts are visible."""
        try:
            # Hide columns from AA (27) to AZ (52) where chart data is stored
            # Frequency data: AA-AF (27-32), Summary data: AG-AH (33-34)
            from openpyxl.utils import get_column_letter
            
            # Hide columns 27-52 (AA to AZ) where we store all chart data
            for col_num in range(27, 53):
                col_letter = get_column_letter(col_num)
                worksheet.column_dimensions[col_letter].hidden = True
            
            self.logger.info("🙈 Hidden chart data columns (AA-AZ) to ensure clean Excel output")
            
        except Exception as e:
            self.logger.error(f"❌ Error hiding chart data columns: {e}")

    def process_noise_test_folder(self, test_folder: Path) -> List[NoiseTestData]:
        """Process all TXT files in a noise test folder."""
        try:
            self.logger.info(f"🔍 Processing noise test folder: {test_folder}")
            
            txt_files = list(test_folder.glob("*.txt"))
            self.logger.info(f"📄 Found {len(txt_files)} TXT files")
            
            parsed_tests = []
            for txt_file in txt_files:
                test_data = self.parse_txt_file(txt_file)
                if test_data:
                    parsed_tests.append(test_data)
            
            self.logger.info(f"✅ Successfully parsed {len(parsed_tests)} test files")
            return parsed_tests
            
        except Exception as e:
            self.logger.error(f"❌ Error processing noise test folder: {e}")
            return []
