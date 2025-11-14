import logging
from typing import List, Dict, Optional, cast
from pathlib import Path

import pandas as pd
import openpyxl
from xlsxwriter.workbook import Workbook as XlsxWorkbook

from ..config.app_config import AppConfig
from ..data.models import NoiseTestInfo, MotorTestData
from .builders.excel_formatter import ExcelFormatter
from .builders.sap_sheet_builder import SapSheetBuilder
from .builders.summary_sheet_builder import SummarySheetBuilder
from .builders.comparison_sheet_builder import ComparisonSheetBuilder
from ..analysis.image_utils import extract_dominant_colors
from ..utils.common import sanitize_sheet_name
from .excel_profiler import ExcelProfiler
from ..services.test_lab_summary_loader import TestLabSummaryLoader

logger = logging.getLogger(__name__)

class ExcelReport:
    """Main class to orchestrate Excel report generation."""

    def __init__(self, config: AppConfig, noise_handler=None):
        self.config = config
        self.logger = logging.getLogger(__class__.__name__)
        self.output_path = Path(config.output_path) if config.output_path else None
        if not self.output_path:
            raise ValueError("Output path must be set in the configuration.")
        
        self.logo_path = config.logo_path
        self.noise_handler = noise_handler  # Store the noise handler
        self.writer: Optional[pd.ExcelWriter] = None
        self.workbook: Optional[XlsxWorkbook] = None
        self.formatter: Optional[ExcelFormatter] = None
        self.logo_tab_colors: List[str] = []
        
        # Performance profiling
        self.profiler = ExcelProfiler("Excel Report Generation")
        self.enable_profiling = False  # Can be enabled via config or env var

    def generate(self, grouped_data: Dict[str, List[MotorTestData]], 
                 all_tests_summary: List[MotorTestData], 
                 all_noise_tests_by_sap: Dict[str, List[NoiseTestInfo]],
                 comparison_data: Dict[str, List[MotorTestData]],
                 multiple_comparisons: Optional[List[Dict]] = None,
                 lf_tests_by_sap: Optional[Dict[str, List]] = None) -> bool:
        """
        Generates the full Excel report with all its components.
        Returns True on success, False on failure.
        
        Args:
            grouped_data: SAP-grouped motor test data
            all_tests_summary: Summary of all tests
            all_noise_tests_by_sap: Noise tests grouped by SAP
            comparison_data: Legacy single comparison data
            multiple_comparisons: New multiple comparison groups
            lf_tests_by_sap: Life Test (LF) data grouped by SAP
        """
        if not self.output_path:
            self.logger.error("Cannot generate report without a valid output path.")
            return False

        # Start profiling session
        if self.enable_profiling:
            self.profiler.start_session()

        try:
            with self.profiler.time_operation("excel_writer_open"):
                writer = pd.ExcelWriter(self.output_path, engine='xlsxwriter')
            
            with writer:
                self.writer = writer
                self.workbook = cast(XlsxWorkbook, writer.book)
                
                with self.profiler.time_operation("extract_logo_colors"):
                    if self.config.logo_path and Path(self.config.logo_path).exists():
                        self.logo_tab_colors = extract_dominant_colors(str(self.config.logo_path))
                    else:
                        self.logger.warning("Logo path not found, using default colors.")
                        self.logo_tab_colors = ['#0070C0', '#C00000', '#00B050', '#FFC000']

                if not self.workbook:
                    self.logger.error("Workbook not created. Aborting.")
                    return False
                
                with self.profiler.time_operation("create_formatter"):
                    self.formatter = ExcelFormatter(self.workbook, self.logo_tab_colors)

                sap_sheet_name_map = {sap: sanitize_sheet_name(f"SAP_{sap}") for sap in grouped_data.keys()}

                # Log LF data status
                if lf_tests_by_sap:
                    total_lf = sum(len(tests) for tests in lf_tests_by_sap.values())
                    self.logger.info(f"📊 ExcelWriter: Received {total_lf} LF test(s) for {len(lf_tests_by_sap)} SAP code(s)")
                else:
                    self.logger.info("📊 ExcelWriter: No LF data received (lf_tests_by_sap is empty or None)")
                
                with self.profiler.time_operation("create_summary_sheet"):
                    self._create_summary_sheet(all_tests_summary, sap_sheet_name_map, multiple_comparisons, lf_tests_by_sap)
                
                with self.profiler.time_operation("create_sap_sheets"):
                    self._create_sap_sheets(grouped_data, all_noise_tests_by_sap)
                
                if self.config.include_comparison:
                    # Create legacy single comparison sheet if comparison_data exists
                    if comparison_data:
                        with self.profiler.time_operation("create_comparison_sheet"):
                            self._create_comparison_sheet(comparison_data, sap_sheet_name_map)
                    
                    # Create multiple comparison sheets if multiple_comparisons exist
                    if multiple_comparisons:
                        with self.profiler.time_operation("create_multiple_comparison_sheets"):
                            self._create_multiple_comparison_sheets(multiple_comparisons, grouped_data, sap_sheet_name_map)

            self.logger.info(f"Successfully generated Excel report at {self.output_path}")
            
            # Copy raw Scheda and Collaudo sheets after xlsxwriter finishes
            with self.profiler.time_operation("copy_raw_test_lab_sheets"):
                self._copy_raw_test_lab_sheets(grouped_data)
            
            # End profiling and print report
            if self.enable_profiling:
                self.profiler.end_session()
                self.profiler.print_report()
            
            return True
        except Exception as e:
            self.logger.error(f"Failed to generate Excel report: {e}", exc_info=True)
            if self.enable_profiling:
                self.profiler.end_session()
            return False

    def _create_summary_sheet(self, all_tests_summary: List[MotorTestData], sap_sheet_name_map: Dict[str, str], multiple_comparisons: Optional[List[Dict]] = None, lf_tests_by_sap: Optional[Dict[str, List]] = None):
        """Creates the main summary sheet."""
        if not self.workbook or not self.formatter: return
        summary_builder = SummarySheetBuilder(
            workbook=self.workbook,
            all_motor_tests=all_tests_summary,
            formatter=self.formatter,
            logo_tab_colors=self.logo_tab_colors,
            sap_sheet_name_map=sap_sheet_name_map,
            multiple_comparisons=multiple_comparisons or [],
            lf_tests_by_sap=lf_tests_by_sap or {}
        )
        summary_builder.build()

    def _create_sap_sheets(self, grouped_data: Dict[str, List[MotorTestData]], all_noise_tests_by_sap: Dict[str, List[NoiseTestInfo]]):
        """Creates a sheet for each SAP code."""
        if not self.workbook or not self.formatter: return
        for sap_code, tests in grouped_data.items():
            sap_builder = SapSheetBuilder(
                workbook=self.workbook,
                sap_code=sap_code,
                motor_tests=tests,
                formatter=self.formatter,
                config=self.config,
                logo_tab_colors=self.logo_tab_colors,
                all_noise_tests=all_noise_tests_by_sap.get(sap_code, []),
                noise_handler=self.noise_handler  # Pass the noise handler
            )
            sap_builder.build()

    def _create_comparison_sheet(self, comparison_data: Dict[str, List[MotorTestData]], sap_sheet_name_map: Dict[str, str]):
        """Creates the comparison sheet if enabled."""
        if not self.workbook or not self.formatter: return
        comp_builder = ComparisonSheetBuilder(
            workbook=self.workbook,
            comparison_data=comparison_data,
            formatter=self.formatter,
            logo_colors=self.logo_tab_colors,
            sap_sheet_name_map=sap_sheet_name_map,
            config=self.config
        )
        comp_builder.build()

    def _create_multiple_comparison_sheets(self, multiple_comparisons: List[Dict], grouped_data: Dict[str, List[MotorTestData]], sap_sheet_name_map: Dict[str, str]):
        """Creates multiple comparison sheets based on user-defined comparison groups."""
        if not self.workbook or not self.formatter: 
            return
            
        self.logger.info(f"Creating {len(multiple_comparisons)} comparison sheets")
        
        for i, comparison_group in enumerate(multiple_comparisons):
            group_name = comparison_group.get('name', f'Comparison {i+1}')
            try:
                # Extract comparison group data
                test_labs = comparison_group.get('test_labs', [])
                description = comparison_group.get('description', '')
                
                self.logger.info(f"Creating comparison sheet '{group_name}' with test labs: {test_labs}")
                
                # Build comparison data for this group
                comparison_data = self._build_comparison_data_for_group(test_labs, grouped_data)
                
                if not comparison_data:
                    self.logger.warning(f"No data found for comparison group '{group_name}', skipping")
                    continue
                
                # Create the comparison sheet with a unique name
                sheet_name = f"Comparison_{i+1}"
                comp_builder = ComparisonSheetBuilder(
                    workbook=self.workbook,
                    comparison_data=comparison_data,
                    formatter=self.formatter,
                    logo_colors=self.logo_tab_colors,
                    sap_sheet_name_map=sap_sheet_name_map,
                    config=self.config,
                    custom_sheet_name=sheet_name,
                    custom_title=group_name,
                    custom_description=description
                )
                comp_builder.build()
                
            except Exception as e:
                self.logger.error(f"Error creating comparison sheet for group '{group_name}': {e}")
    
    def _build_comparison_data_for_group(self, test_labs: List[str], grouped_data: Dict[str, List[MotorTestData]]) -> Dict[str, List[MotorTestData]]:
        """Build comparison data for a specific group of test labs."""
        comparison_data = {}
        
        # Search through all SAP data to find the specified test labs
        for sap_code, tests in grouped_data.items():
            matching_tests = []
            for test in tests:
                if test.test_number in test_labs:
                    matching_tests.append(test)
            
            if matching_tests:
                comparison_data[sap_code] = matching_tests
        
        return comparison_data

    def _copy_raw_test_lab_sheets(self, grouped_data: Dict[str, List[MotorTestData]]) -> None:
        """Copy raw Scheda and Collaudo sheets from test lab workbooks after report generation."""
        if not self.output_path or not self.output_path.exists():
            self.logger.warning("Cannot copy raw sheets: output file not found")
            return
        
        # Initialize the test lab summary loader
        test_lab_root = self.config.test_lab_root
        if not test_lab_root or not Path(test_lab_root).exists():
            self.logger.warning("Test lab root path not configured or not found, skipping raw sheet copying")
            return
        
        loader = TestLabSummaryLoader(base_path=Path(test_lab_root))
        
        # Open the generated report with openpyxl
        try:
            openpyxl_wb = openpyxl.load_workbook(self.output_path)
            
            # Process each SAP code
            for sap_code, tests in grouped_data.items():
                self._create_collaudo_nominale_sheet(openpyxl_wb, sap_code, tests, loader)
            
            # Save the modified workbook
            openpyxl_wb.save(self.output_path)
            openpyxl_wb.close()
            
            self.logger.info("Successfully created COLLAUDO NOMINALE sheets")
            
        except Exception as e:
            self.logger.error(f"Failed to create COLLAUDO NOMINALE sheets: {e}", exc_info=True)
    
    def _create_collaudo_nominale_sheet(
        self, 
        workbook: openpyxl.Workbook, 
        sap_code: str, 
        tests: List[MotorTestData],
        loader: 'TestLabSummaryLoader'
    ) -> None:
        """Create a single COLLAUDO NOMINALE sheet for a SAP code with Scheda on left, Collaudo on right."""
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as OpenpyxlImage
        import openpyxl.styles
        
        sheet_name = f"COLLAUDO NOMINALE - {sap_code}"
        
        # Ensure unique sheet name (max 31 chars for Excel)
        if len(sheet_name) > 31:
            sheet_name = f"COLLAUDO NOM - {sap_code}"[:31]
        
        if sheet_name in workbook.sheetnames:
            counter = 1
            while f"{sheet_name}_{counter}" in workbook.sheetnames:
                counter += 1
            sheet_name = f"{sheet_name}_{counter}"[:31]
        
        # Create sheet
        target_sheet = workbook.create_sheet(title=sheet_name)
        
        # Set tab color to match theme (use logo colors if available)
        try:
            if self.logo_tab_colors and len(self.logo_tab_colors) > 0:
                # Use first logo color for consistency with SAP sheets
                color_hex = self.logo_tab_colors[0].lstrip('#')
                target_sheet.sheet_properties.tabColor = color_hex
            else:
                # Default blue color
                target_sheet.sheet_properties.tabColor = "0070C0"
        except:
            pass
        
        # Add logo at top (matching other sheets)
        logo_added = False
        if self.config.logo_path and Path(self.config.logo_path).exists():
            try:
                img = OpenpyxlImage(str(self.config.logo_path))
                # Scale logo to fit in header area
                img.width = 120
                img.height = 60
                target_sheet.add_image(img, 'A1')
                logo_added = True
            except Exception as e:
                self.logger.warning(f"Could not add logo to COLLAUDO NOMINALE sheet: {e}")
        
        # Collect all Scheda and Collaudo sheets from all tests for this SAP
        all_scheda_data = []
        all_collaudo_data = []
        workbooks_to_close = []
        
        for test in tests:
            test_number = test.test_number
            if not test_number:
                continue
            
            try:
                sheets_data = loader.get_raw_sheets_data(test_number)
                if sheets_data:
                    scheda_sheets = sheets_data.get('scheda', [])
                    collaudo_sheets = sheets_data.get('collaudo', [])
                    source_wb = sheets_data.get('workbook')
                    
                    for sheet in scheda_sheets:
                        all_scheda_data.append((test_number, sheet))
                    for sheet in collaudo_sheets:
                        all_collaudo_data.append((test_number, sheet))
                    
                    if source_wb:
                        workbooks_to_close.append(source_wb)
                        
            except Exception as e:
                self.logger.error(f"Failed to get sheets for test {test_number}: {e}")
        
        if not all_scheda_data and not all_collaudo_data:
            self.logger.warning(f"No Scheda or Collaudo sheets found for SAP {sap_code}")
            workbook.remove(target_sheet)
            for wb in workbooks_to_close:
                wb.close()
            return
        
        # Set column widths for header area
        for col in range(1, 17):  # A-P
            target_sheet.column_dimensions[get_column_letter(col)].width = 15
        
        # Add header section (start after logo if present)
        header_start_row = 4 if logo_added else 1
        
        header_font = openpyxl.styles.Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        
        # Main title spanning both sections
        target_sheet.cell(row=header_start_row, column=1, value=f"COLLAUDO NOMINALE - SAP {sap_code}")
        target_sheet.cell(row=header_start_row, column=1).font = header_font
        target_sheet.cell(row=header_start_row, column=1).fill = header_fill
        target_sheet.cell(row=header_start_row, column=1).alignment = header_alignment
        target_sheet.row_dimensions[header_start_row].height = 25
        
        # Determine layout columns
        max_scheda_cols = 0
        for _, sheet in all_scheda_data:
            max_col = sheet.max_column or 10
            max_scheda_cols = max(max_scheda_cols, max_col)
        
        # Merge title across full width
        right_col_start = max_scheda_cols + 4  # 3-column gap
        max_collaudo_cols = 0
        for _, sheet in all_collaudo_data:
            max_col = sheet.max_column or 10
            max_collaudo_cols = max(max_collaudo_cols, max_col)
        
        total_cols = right_col_start + max_collaudo_cols
        try:
            target_sheet.merge_cells(start_row=header_start_row, start_column=1, end_row=header_start_row, end_column=min(total_cols, 50))
        except:
            pass
        
        # Section headers
        section_font = openpyxl.styles.Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        section_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        
        section_header_row = header_start_row + 1
        
        # Scheda header
        target_sheet.cell(row=section_header_row, column=1, value="SCHEDA / SCHEDA SR")
        target_sheet.cell(row=section_header_row, column=1).font = section_font
        target_sheet.cell(row=section_header_row, column=1).fill = section_fill
        target_sheet.cell(row=section_header_row, column=1).alignment = section_alignment
        target_sheet.row_dimensions[section_header_row].height = 20
        try:
            target_sheet.merge_cells(start_row=section_header_row, start_column=1, end_row=section_header_row, end_column=max_scheda_cols)
        except:
            pass
        
        # Collaudo header
        target_sheet.cell(row=section_header_row, column=right_col_start, value="COLLAUDO / COLLAUDO SR")
        target_sheet.cell(row=section_header_row, column=right_col_start).font = section_font
        target_sheet.cell(row=section_header_row, column=right_col_start).fill = section_fill
        target_sheet.cell(row=section_header_row, column=right_col_start).alignment = section_alignment
        try:
            target_sheet.merge_cells(start_row=section_header_row, start_column=right_col_start, end_row=section_header_row, end_column=right_col_start + max_collaudo_cols - 1)
        except:
            pass
        
        # Layout: Scheda on left, Collaudo on right
        current_row_left = header_start_row + 3  # Start after headers
        current_row_right = header_start_row + 3
        left_col = 1
        
        # Test label formatting
        test_label_font = openpyxl.styles.Font(name='Calibri', size=11, bold=True, color="1F4E78")
        test_label_fill = openpyxl.styles.PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        test_label_alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
        
        # Copy Scheda sheets vertically on the left
        for test_number, source_sheet in all_scheda_data:
            # Add test label with styling
            label_cell = target_sheet.cell(row=current_row_left, column=left_col, value=f"Test {test_number} - {source_sheet.title}")
            label_cell.font = test_label_font
            label_cell.fill = test_label_fill
            label_cell.alignment = test_label_alignment
            target_sheet.row_dimensions[current_row_left].height = 20
            
            # Merge label across Scheda section
            try:
                target_sheet.merge_cells(start_row=current_row_left, start_column=left_col, end_row=current_row_left, end_column=max_scheda_cols)
            except:
                pass
            
            current_row_left += 1
            
            # Copy sheet content
            self._copy_sheet_content_to_position(source_sheet, target_sheet, current_row_left, left_col)
            
            # Update current row (move down past this sheet + gap)
            rows_copied = source_sheet.max_row or 20
            current_row_left += rows_copied + 3  # 3-row gap between sheets
        
        # Copy Collaudo sheets vertically on the right
        for test_number, source_sheet in all_collaudo_data:
            # Add test label with styling
            label_cell = target_sheet.cell(row=current_row_right, column=right_col_start, value=f"Test {test_number} - {source_sheet.title}")
            label_cell.font = test_label_font
            label_cell.fill = test_label_fill
            label_cell.alignment = test_label_alignment
            target_sheet.row_dimensions[current_row_right].height = 20
            
            # Merge label across Collaudo section
            try:
                target_sheet.merge_cells(start_row=current_row_right, start_column=right_col_start, end_row=current_row_right, end_column=right_col_start + max_collaudo_cols - 1)
            except:
                pass
            
            current_row_right += 1
            
            # Copy sheet content
            self._copy_sheet_content_to_position(source_sheet, target_sheet, current_row_right, right_col_start)
            
            # Update current row
            rows_copied = source_sheet.max_row or 20
            current_row_right += rows_copied + 3
        
        # Set print settings
        try:
            target_sheet.page_setup.orientation = target_sheet.ORIENTATION_LANDSCAPE
            target_sheet.page_setup.paperSize = target_sheet.PAPERSIZE_A4
            target_sheet.page_setup.fitToPage = True
            target_sheet.page_setup.fitToHeight = 0  # Fit to width only
            target_sheet.page_setup.fitToWidth = 1
        except:
            pass
        
        # Close source workbooks
        for wb in workbooks_to_close:
            try:
                wb.close()
            except:
                pass
        
        self.logger.info(f"Created sheet '{sheet_name}' with {len(all_scheda_data)} Scheda and {len(all_collaudo_data)} Collaudo sheets")
    
    def _copy_sheet_content_to_position(
        self,
        source_sheet: 'openpyxl.worksheet.worksheet.Worksheet',
        target_sheet: 'openpyxl.worksheet.worksheet.Worksheet',
        start_row: int,
        start_col: int
    ) -> None:
        """Copy worksheet content to a specific position in target sheet with full formatting and formulas."""
        from copy import copy as copy_obj, deepcopy
        import openpyxl.styles
        from openpyxl.styles import Color, PatternFill, Border, Side, Font, Alignment, Protection
        
        # Copy column widths (offset by start_col)
        for col_letter, col_dim in source_sheet.column_dimensions.items():
            try:
                col_idx = openpyxl.utils.column_index_from_string(col_letter)
                target_col_letter = openpyxl.utils.get_column_letter(col_idx + start_col - 1)
                target_sheet.column_dimensions[target_col_letter].width = col_dim.width
                # Copy hidden state
                if col_dim.hidden:
                    target_sheet.column_dimensions[target_col_letter].hidden = True
            except Exception as e:
                self.logger.debug(f"Column dimension copy failed: {e}")
        
        # Copy row heights and hidden states (offset by start_row)
        for row_num, row_dim in source_sheet.row_dimensions.items():
            try:
                target_sheet.row_dimensions[row_num + start_row - 1].height = row_dim.height
                if row_dim.hidden:
                    target_sheet.row_dimensions[row_num + start_row - 1].hidden = True
            except Exception as e:
                self.logger.debug(f"Row dimension copy failed: {e}")
        
        # Copy cells with FULL formatting and formulas
        for row in source_sheet.iter_rows():
            for source_cell in row:
                if source_cell.row is None or source_cell.column is None:
                    continue
                
                target_row = source_cell.row + start_row - 1
                target_col = source_cell.column + start_col - 1
                target_cell = target_sheet.cell(row=target_row, column=target_col)
                
                # Copy cell value based on type
                try:
                    if source_cell.data_type == 'f':  # Formula
                        target_cell.value = source_cell.value
                        target_cell.data_type = 'f'
                    else:
                        target_cell.value = source_cell.value
                        if source_cell.data_type:
                            target_cell.data_type = source_cell.data_type
                except Exception as e:
                    self.logger.debug(f"Value copy failed for {source_cell.coordinate}: {e}")
                    try:
                        target_cell.value = source_cell.value
                    except:
                        pass
                
                # Copy ALL formatting attributes - use direct assignment
                if source_cell.has_style:
                    try:
                        # Direct copy approach - more reliable than copy()
                        target_cell._style = source_cell._style
                        
                    except Exception as e:
                        # Fallback: manual attribute copy
                        try:
                            if source_cell.font:
                                target_cell.font = source_cell.font
                            if source_cell.fill:
                                target_cell.fill = source_cell.fill
                            if source_cell.border:
                                target_cell.border = source_cell.border
                            if source_cell.number_format:
                                target_cell.number_format = source_cell.number_format
                            if source_cell.alignment:
                                target_cell.alignment = source_cell.alignment
                            if source_cell.protection:
                                target_cell.protection = source_cell.protection
                        except Exception as e2:
                            self.logger.debug(f"Style copy failed for {source_cell.coordinate}: {e2}")
        
        # Copy merged cells (adjust coordinates)
        for merged_range in source_sheet.merged_cells.ranges:
            min_row = merged_range.min_row + start_row - 1
            max_row = merged_range.max_row + start_row - 1
            min_col = merged_range.min_col + start_col - 1
            max_col = merged_range.max_col + start_col - 1
            
            try:
                target_sheet.merge_cells(
                    start_row=min_row,
                    start_column=min_col,
                    end_row=max_row,
                    end_column=max_col
                )
            except Exception as e:
                self.logger.debug(f"Merge failed for range {merged_range}: {e}")
        
        # Copy conditional formatting (if any)
        try:
            if hasattr(source_sheet, 'conditional_formatting') and source_sheet.conditional_formatting:
                for cf_range, cf_rules in source_sheet.conditional_formatting._cf_rules.items():
                    # Adjust range for offset
                    # Note: This is complex and may not work perfectly for all cases
                    pass
        except Exception as e:
            self.logger.debug(f"Conditional formatting copy skipped: {e}")

