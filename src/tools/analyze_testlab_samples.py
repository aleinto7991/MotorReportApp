"""Scan multiple test-lab Excel workbooks and extract key fields.

This script will:
- Copy up to N recent .xlsx files from the TEST_LAB path into `logs/` (only .xlsx to avoid .xls engine issues)
- Open each workbook with openpyxl and look for sheets named 'Scheda SR' and 'Collaudo SR'
- Fallback: scan all sheets for keyword hints like 'Prova', 'Codice SAP', 'MEDIA', 'SIGMA'
- Extract: test number, date, SAP code, and summary rows (MEDIA/SIGMA) when present
- Produce a JSON summary at `logs/analysis_report.json` and print brief results

Adjust SOURCE_ROOT to your environment if needed.
"""
from pathlib import Path
import shutil
import json
import re
from datetime import datetime, timezone
import openpyxl


# Configuration - change if needed
SOURCE_ROOT = Path(r"C:\Users\aintorbida\OneDrive - AMETEK Inc\ENG & Quality\UTE_wrk\LAB\TEST_LAB\CARICHI NOMINALI")
TARGET_LOGS = Path(__file__).resolve().parents[2] / "logs"
MAX_FILES = 20

TARGET_LOGS.mkdir(parents=True, exist_ok=True)


def find_source_xlsx_files(limit=MAX_FILES):
    """Find recent .xlsx files across year folders (searches depth 2)."""
    files = []
    if not SOURCE_ROOT.exists():
        return files
    # Look in year folders first (directories that look like 4-digit years)
    for year_dir in sorted([p for p in SOURCE_ROOT.iterdir() if p.is_dir()], reverse=True):
        # collect .xlsx files (ignore everything else)
        for f in sorted(year_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True):
            files.append(f)
            if len(files) >= limit:
                return files
    # fallback: also look at top-level .xlsx
    for f in sorted(SOURCE_ROOT.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True):
        if f not in files:
            files.append(f)
            if len(files) >= limit:
                return files
    return files


def copy_samples(files):
    copied = []
    for src in files:
        dst = TARGET_LOGS / src.name
        try:
            shutil.copy2(src, dst)
            copied.append(dst)
        except Exception as e:
            print(f"Failed to copy {src}: {e}")
    return copied


def text_norm(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def find_right_value_cell(ws, row: int, col: int, max_offset: int = 6):
    """Return the first non-empty cell to the right of (row, col)."""
    for offset in range(1, max_offset + 1):
        c = ws.cell(row=row, column=col + offset)
        if c.value is not None and text_norm(c.value) != "":
            return c
    return None


def extract_from_workbook(path: Path):
    info = {"file": str(path), "sheets": {}, "found": False}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        info["error"] = str(e)
        return info

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sheet_info = {"rows_scanned": 0, "found_fields": {}}

        # scan first 80 rows x 20 cols for hints
        for r in ws.iter_rows(min_row=1, max_row=80, min_col=1, max_col=20, values_only=False):
            sheet_info["rows_scanned"] += 1
            for cell in r:
                val = cell.value
                if val is None:
                    continue
                v = text_norm(val)
                # detect test number line
                if re.search(r"prova n|prova n\W|prova numero|prova\s*:\s*", v, re.I):
                    if "test_number" not in sheet_info["found_fields"]:
                        row_idx = cell.row or 0
                        col_idx = cell.column or 0
                        target = find_right_value_cell(ws, row_idx, col_idx) if row_idx and col_idx else None
                        if target:
                            sheet_info["found_fields"]["test_number"] = {
                                "value_cell": target.coordinate,
                                "value": text_norm(target.value)
                            }
                if re.search(r"cod(ice)?\s*sap", v, re.I):
                    if "sap_code" not in sheet_info["found_fields"]:
                        row_idx = cell.row or 0
                        col_idx = cell.column or 0
                        target = find_right_value_cell(ws, row_idx, col_idx) if row_idx and col_idx else None
                        if target:
                            sheet_info["found_fields"]["sap_code"] = {
                                "value_cell": target.coordinate,
                                "value": text_norm(target.value)
                            }
                if re.search(r"\bdata\b|date", v, re.I):
                    if "date" not in sheet_info["found_fields"]:
                        row_idx = cell.row or 0
                        col_idx = cell.column or 0
                        target = find_right_value_cell(ws, row_idx, col_idx) if row_idx and col_idx else None
                        if target:
                            raw = target.value
                            if isinstance(raw, datetime):
                                formatted = raw.date().isoformat()
                            else:
                                formatted = text_norm(raw)
                            sheet_info["found_fields"]["date"] = {
                                "value_cell": target.coordinate,
                                "value": formatted
                            }
                if re.search(r"media|sigma", v, re.I):
                    # capture the row values for summary rows
                    row_idx = cell.row or 0
                    if not row_idx:
                        continue
                    row_vals = [text_norm(c.value) for c in ws[row_idx]]
                    sheet_info["found_fields"].setdefault("summary_rows", {})
                    key = v.upper()
                    sheet_info["found_fields"]["summary_rows"][key] = {"row": cell.row, "values": row_vals}

        # Also detect if sheet name is exactly our target
        if sheet.lower() in ("scheda sr", "scheda_sr"):
            sheet_info["sheet_type"] = "scheda"
        if sheet.lower() in ("collaudo sr", "collaudo_sr"):
            sheet_info["sheet_type"] = "collaudo"

        if sheet_info["found_fields"]:
            info["found"] = True
        info["sheets"][sheet] = sheet_info

    return info


def main():
    samples = find_source_xlsx_files()
    print(f"Found {len(samples)} source .xlsx files to copy")
    copied = copy_samples(samples)
    print(f"Copied {len(copied)} files into {TARGET_LOGS}")

    results = []
    for f in copied:
        print(f"Analyzing {f.name}...")
        try:
            res = extract_from_workbook(f)
            results.append(res)
        except Exception as e:
            results.append({"file": str(f), "error": str(e)})

    out = TARGET_LOGS / "analysis_report.json"
    with open(out, "w", encoding="utf-8") as fh:
        json.dump({"generated_at": datetime.now(timezone.utc).isoformat(), "results": results}, fh, indent=2, ensure_ascii=False)

    print(f"Analysis written to: {out}")


if __name__ == "__main__":
    main()
