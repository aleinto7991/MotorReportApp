"""Utilities for locating and extracting summary data from test-lab Excel workbooks.

This module encapsulates the logic required to locate the so-called "A" test
workbooks stored in the TEST_LAB/"CARICHI NOMINALI" folder hierarchy and to
collect the specific summary blocks required for the SAP report.
"""
from __future__ import annotations

import logging
import re
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Dict, Iterable, List, Optional, Tuple, Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string

from ..data.models import CollaudoSummary, SchedaSummary, TestLabSummary

logger = logging.getLogger(__name__)


@dataclass
class _WorkbookExtractionResult:
    """Internal helper bundle returned by the extractor."""

    scheda: Optional[SchedaSummary]
    collaudo: Optional[CollaudoSummary]


@dataclass(frozen=True)
class TestLabWorkbookMatch:
    """Describe a workbook located within the Test-Lab hierarchy."""

    requested_test_number: str
    matched_test_number: str
    match_strategy: str
    path: Path
    year_folder: Optional[str]


class TestLabSummaryLoader:
    """Locate and extract summary data from test-lab workbooks."""

    __test__ = False  # Explicitly opt-out of pytest class discovery

    # Order used when writing rows in the SAP sheet
    _SCHEda_HEADERS: List[str] = [
        "Orifice (mm)",
        "Watt",
        "Watt c.",
        "mmH2O",
        "mmH2O c.",
        "Portata",
        "Air Watt",
        "Eff.%",
    ]

    # Mapping of sanitized header text to canonical header names
    _SCHEda_HEADER_ALIASES: Dict[str, str] = {
        "orifice": "Orifice (mm)",
        "orificemm": "Orifice (mm)",
        "watt": "Watt",
        "wattc": "Watt c.",
        "mmh2o": "mmH2O",
        "mmh2oc": "mmH2O c.",
        "portata": "Portata",
        "airwatt": "Air Watt",
        "eff": "Eff.%",
        "eff%": "Eff.%",
    }

    _COLLAUDO_COLUMNS: List[tuple[str, int]] = [
        ("Ampere 22.2", 2),
        ("Ampere BA", 3),
        ("Ampere BC", 4),
        ("Watt 22.2", 5),
        ("Watt BA", 6),
        ("Watt BC", 7),
        ("RPM 22.2", 8),
        ("RPM BA", 9),
        ("RPM BC", 10),
        ("mmH2O 22.2", 11),
        ("mmH2O BA", 12),
        ("mmH2O BC", 13),
    ]

    _COLLAUDO_HEADER_ALIASES: Dict[str, str] = {
        "ampere222": "Ampere 22.2",
        "ampere22": "Ampere 22.2",
        "ampereba": "Ampere BA",
        "amperebc": "Ampere BC",
        "watt222": "Watt 22.2",
        "watt22": "Watt 22.2",
        "wattba": "Watt BA",
        "wattbc": "Watt BC",
        "rpm222": "RPM 22.2",
        "rpm22": "RPM 22.2",
        "rpmba": "RPM BA",
        "rpmbc": "RPM BC",
        "mmh2o222": "mmH2O 22.2",
        "mmh2o22": "mmH2O 22.2",
        "mmh2oba": "mmH2O BA",
        "mmh2obc": "mmH2O BC",
    }

    def __init__(self, base_path: Optional[str], logger_: Optional[logging.Logger] = None) -> None:
        self.base_path = Path(base_path) if base_path else None
        self.logger = logger_ or logger

    # ------------------------------------------------------------------
    @property
    def available(self) -> bool:
        return bool(self.base_path and self.base_path.exists())

    # ------------------------------------------------------------------
    def load_summary(self, test_number: str, override_path: Optional[str] = None) -> Optional[TestLabSummary]:
        """Return the extracted summary for *test_number* or ``None`` if not found."""
        workbook_path = None
        matched_stem = test_number
        match_strategy = "auto"

        if override_path:
            workbook_path = Path(override_path)
            if not workbook_path.exists():
                self.logger.warning("Override path provided but does not exist: %s", override_path)
                return None
            match_strategy = "manual_override"
            self.logger.info("Using manual override for test %s: %s", test_number, workbook_path)
        else:
            if not self.available:
                self.logger.debug("Test lab base directory not configured; skipping lookup")
                return None

            self.logger.debug(
                "Starting summary lookup for test %s (base directory: %s)",
                test_number,
                self.base_path,
            )

            match = self.locate_workbook(test_number)
            if not match:
                self.logger.info("No test-lab workbook found for %s", test_number)
                return None

            workbook_path = match.path
            matched_stem = match.matched_test_number
            match_strategy = match.match_strategy

        try:
            extraction = self._extract_from_workbook(workbook_path)
            raw_sheets = self._extract_raw_sheets(workbook_path)
        except Exception as exc:  # pragma: no cover - guard against unexpected formats
            self.logger.warning("Failed to parse test-lab workbook %s: %s", workbook_path, exc)
            return None

        if not extraction.scheda and not extraction.collaudo and not raw_sheets:
            self.logger.info(
                "Workbook %s does not contain expected Scheda/Collaudo summaries", workbook_path
            )
            return None

        return TestLabSummary(
            source_path=str(workbook_path),
            scheda=extraction.scheda,
            collaudo_media=extraction.collaudo,
            matched_test_number=matched_stem,
            match_strategy=match_strategy,
            raw_sheets=raw_sheets,
        )

    # ------------------------------------------------------------------
    def locate_workbook(self, test_number: str) -> Optional[TestLabWorkbookMatch]:
        """Return details about the workbook backing ``test_number`` if present."""
        if not self.available:
            self.logger.debug("Test lab base directory not configured; skipping lookup")
            return None

        match = self._locate_workbook(test_number)
        if not match:
            return None

        workbook_path, matched_stem, match_strategy = match
        year_folder = self._derive_year_folder(workbook_path)
        return TestLabWorkbookMatch(
            requested_test_number=test_number,
            matched_test_number=matched_stem,
            match_strategy=match_strategy,
            path=workbook_path,
            year_folder=year_folder,
        )

    def _locate_workbook(self, test_number: str) -> Optional[Tuple[Path, str, str]]:
        assert self.base_path is not None  # guarded by available property

        normalized = re.sub(r"[^0-9A-Za-z]+", "", test_number or "").upper()
        if not normalized:
            return None

        requested_is_alias = normalized.endswith("A")
        primary_candidates: List[str] = [normalized]
        fallback_candidates: List[str] = []

        # Strict matching requested: do not fallback to base or 'A' variant
        # if requested_is_alias:
        #     base_candidate = normalized.rstrip("A")
        #     if base_candidate:
        #         fallback_candidates.append(base_candidate)
        # else:
        #     fallback_candidates.append(f"{normalized}A")

        search_dirs = list(self._iter_search_directories())

        primary_allow_prefix = requested_is_alias
        primary_match = self._search_candidates(search_dirs, primary_candidates, allow_prefix=primary_allow_prefix)
        if primary_match:
            path, candidate, prefix_used = primary_match
            strategy = "prefix" if prefix_used else "exact"
            return path, candidate, strategy

        fallback_match = self._search_candidates(search_dirs, fallback_candidates, allow_prefix=True)
        if fallback_match:
            path, candidate, prefix_used = fallback_match
            strategy = "fallback_prefix" if prefix_used else "fallback_exact"
            return path, candidate, strategy

        # if not primary_allow_prefix:
        #     prefix_match = self._search_candidates(search_dirs, primary_candidates, allow_prefix=True)
        #     if prefix_match:
        #         path, candidate, prefix_used = prefix_match
        #         strategy = "prefix" if prefix_used else "exact"
        #         return path, candidate, strategy

        evaluated = primary_candidates + fallback_candidates
        self.logger.info(
            "Could not locate workbook for %s after searching %d directories. "
            "Candidates evaluated: %s. Check that the file exists in the TEST_LAB directory.",
            normalized,
            len(search_dirs),
            evaluated,
        )
        return None

    def _derive_year_folder(self, workbook_path: Path) -> Optional[str]:
        if not self.base_path:
            return None
        try:
            relative_parent = workbook_path.parent.relative_to(self.base_path)
        except ValueError:
            return None
        parts = relative_parent.parts
        if parts:
            return parts[0]
        return None

    def _search_candidates(
        self,
        search_dirs: Iterable[Path],
        stems: List[str],
        *,
        allow_prefix: bool,
    ) -> Optional[Tuple[Path, str, bool]]:
        for stem in stems:
            if not stem:
                continue
            for directory in search_dirs:
                self.logger.debug("Looking for %s.xlsx in %s", stem, directory)
                match = self._find_in_directory(directory, stem, allow_prefix=allow_prefix)
                if match:
                    path, prefix_used = match
                    self.logger.debug("Located workbook %s for candidate %s", path, stem)
                    return path, stem, prefix_used
        return None

    # ------------------------------------------------------------------
    def _iter_search_directories(self) -> Iterable[Path]:
        """Yield directories to search, prioritising year folders in descending order."""
        assert self.base_path is not None

        subdirs = [p for p in self.base_path.iterdir() if p.is_dir()]
        self.logger.debug(
            "Found %d subdirectories in %s: %s",
            len(subdirs),
            self.base_path,
            [p.name for p in subdirs] if subdirs else "(none)",
        )
        # Sort descending by name so the most recent year is checked first
        for folder in sorted(subdirs, key=lambda p: p.name, reverse=True):
            self.logger.debug("Searching in subdirectory: %s", folder)
            yield folder
        self.logger.debug("Searching in base directory: %s", self.base_path)
        yield self.base_path

    # ------------------------------------------------------------------
    def _find_in_directory(
        self,
        directory: Path,
        stem: str,
        *,
        allow_prefix: bool,
    ) -> Optional[Tuple[Path, bool]]:
        target_lower = stem.lower()

        for suffix in (".xlsx", ".xls"):
            exact_path = directory / f"{stem}{suffix}"
            if exact_path.exists():
                return exact_path, False

        workbook_candidates = [
            candidate
            for candidate in directory.iterdir()
            if candidate.suffix.lower() in {".xlsx", ".xls"}
        ]
        
        # Log what files were actually found in this directory
        if workbook_candidates:
            file_list = [c.name for c in workbook_candidates[:15]]  # Show first 15
            if len(workbook_candidates) > 15:
                file_list.append(f"... and {len(workbook_candidates) - 15} more")
            self.logger.debug(
                "Found %d Excel file(s) in %s: %s",
                len(workbook_candidates),
                directory.name,
                file_list,
            )
        else:
            self.logger.debug("No Excel files found in %s", directory.name)

        for candidate in workbook_candidates:
            if candidate.stem.lower() == target_lower:
                return candidate, False

        if not allow_prefix:
            return None

        starts_with_matches = [
            candidate
            for candidate in workbook_candidates
            if candidate.stem.lower().startswith(target_lower)
        ]
        if starts_with_matches:
            starts_with_matches.sort(key=lambda p: self._prefix_sort_key(stem, p))
            return starts_with_matches[0], True

        return None

    def _prefix_sort_key(self, requested_stem: str, candidate: Path) -> Tuple[int, int, float]:
        requested_alias = requested_stem.upper().endswith("A")
        candidate_alias = candidate.stem.upper().endswith("A")
        alias_penalty = 0 if requested_alias == candidate_alias else 1
        length_penalty = abs(len(candidate.stem) - len(requested_stem))
        try:
            mtime_penalty = -candidate.stat().st_mtime
        except OSError:
            mtime_penalty = 0.0
        return alias_penalty, length_penalty, mtime_penalty

    # ------------------------------------------------------------------
    def _extract_from_workbook(self, workbook_path: Path) -> _WorkbookExtractionResult:
        self.logger.debug("Opening workbook %s", workbook_path)

        temp_converted_path: Optional[Path] = None
        workbook_to_read = workbook_path

        if workbook_path.suffix.lower() == ".xls":
            temp_converted_path = self._convert_xls_to_xlsx(workbook_path)
            if not temp_converted_path:
                self.logger.warning(
                    "Skipping workbook %s because legacy XLS conversion failed",
                    workbook_path,
                )
                return _WorkbookExtractionResult(scheda=None, collaudo=None)
            workbook_to_read = temp_converted_path

        wb = None
        try:
            wb = openpyxl.load_workbook(workbook_to_read, data_only=True)
            scheda = self._extract_scheda_summary(wb)
            collaudo = self._extract_collaudo_media(wb)
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
            if temp_converted_path and temp_converted_path.exists():
                temp_converted_path.unlink(missing_ok=True)

        self.logger.debug(
            "Extraction summary for %s -> scheda=%s collaudo=%s",
            workbook_path,
            "yes" if scheda else "no",
            "yes" if collaudo else "no",
        )
        return _WorkbookExtractionResult(scheda=scheda, collaudo=collaudo)

    # ------------------------------------------------------------------
    def _convert_xls_to_xlsx(self, workbook_path: Path) -> Optional[Path]:
        try:
            import pandas as pd
        except ImportError:
            self.logger.warning(
                "Cannot process legacy workbook %s: pandas is required for conversion",
                workbook_path,
            )
            return None

        try:
            with pd.ExcelFile(workbook_path, engine="xlrd") as xls:
                with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    tmp_path = Path(tmp.name)
                with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                    for sheet_name in xls.sheet_names:
                        df = xls.parse(sheet_name, header=None)
                        writer_sheet_name = str(sheet_name)
                        df.to_excel(writer, sheet_name=writer_sheet_name, index=False, header=False)
            return tmp_path
        except ImportError as exc:
            self.logger.warning(
                "Cannot process legacy workbook %s: xlrd is required (%s)",
                workbook_path,
                exc,
            )
        except Exception as exc:  # pragma: no cover - defensive guard
            self.logger.warning(
                "Failed to convert legacy workbook %s: %s",
                workbook_path,
                exc,
            )
        return None

    # ------------------------------------------------------------------
    def _extract_scheda_summary(self, wb) -> Optional[SchedaSummary]:
        self.logger.debug("Attempting to extract Scheda SR block")
        # Be tolerant to sheet name variations similar to Collaudo handling
        scheda_sheet_name = None
        for s in wb.sheetnames:
            if "scheda" in self._normalize_text(s):
                scheda_sheet_name = s
                break
        if not scheda_sheet_name:
            self.logger.debug("Workbook missing any sheet with name containing 'Scheda'")
            return None

        ws: Worksheet = wb[scheda_sheet_name]
        header_candidates = self._find_all_scheda_header_rows(ws)
        if not header_candidates:
            self.logger.debug("Scheda SR headers not detected")
            return None

        # Build a fast list of header rows sorted ascending by row index
        header_candidates.sort(key=lambda x: x[0])

        # 1) Label-first search across the sheet: media → min → max
        max_rows = ws.max_row or 0
        max_cols = min(ws.max_column or 0, 80)
        label_hits: List[tuple[int, str]] = []
        for row_idx in range(1, max_rows + 1):
            row_iter = ws.iter_rows(min_row=row_idx, max_row=row_idx, max_col=max_cols)
            row_cells = next(row_iter, tuple())
            label = self._detect_summary_label(row_cells)
            if label in {"media", "min", "max"}:
                label_hits.append((row_idx, label))

        if not label_hits:
            self.logger.debug("No Media/Min/Max labels detected anywhere in Scheda sheet")
            return None

        # 2) For each hit, pick nearest header row above it, then extract values
        # Group potential rows by their header anchor
        grouped: Dict[int, Dict[str, Dict[str, Optional[float]]]] = {}
        headers_by_anchor: Dict[int, List[str]] = {}
        mapping_by_anchor: Dict[int, Dict[str, int]] = {}
        rowindex_by_anchor: Dict[int, Dict[str, int]] = {}

        def nearest_header_above(r: int) -> Optional[tuple[int, Dict[str, int]]]:
            prior = [h for h in header_candidates if h[0] < r]
            if not prior:
                return None
            return prior[-1]

        for row_idx, label in label_hits:
            anchor = nearest_header_above(row_idx)
            if not anchor:
                continue
            header_row_idx, header_map = anchor
            headers = [h for h in self._SCHEda_HEADERS if h in header_map]
            if not headers:
                continue
            # Extract row values at this row using header_map columns
            row_values: Dict[str, Optional[float]] = {}
            for h in headers:
                col_idx = header_map[h]
                val = ws.cell(row=row_idx, column=col_idx).value
                row_values[h] = self._parse_number(val)

            # Initialize cluster for this anchor if needed
            grouped.setdefault(header_row_idx, {})
            headers_by_anchor.setdefault(header_row_idx, headers)
            mapping_by_anchor.setdefault(header_row_idx, header_map)
            rowindex_by_anchor.setdefault(header_row_idx, {})
            canonical_label = label.capitalize()
            grouped[header_row_idx][canonical_label] = row_values
            rowindex_by_anchor[header_row_idx][canonical_label] = row_idx

        if not grouped:
            self.logger.debug("Labels found but no usable header anchor above them")
            return None

        # 3) Strict selection: require all three labels present and in order Media < Min < Max
        candidate_anchors: List[int] = []
        for ar, rows_map in grouped.items():
            idx_map = rowindex_by_anchor.get(ar, {})
            if not ({"Media", "Min", "Max"} <= set(rows_map.keys())):
                continue
            r_media = idx_map.get("Media")
            r_min = idx_map.get("Min")
            r_max = idx_map.get("Max")
            if not (
                isinstance(r_media, int)
                and isinstance(r_min, int)
                and isinstance(r_max, int)
                and r_media < r_min < r_max
            ):
                continue
            candidate_anchors.append(ar)

        if not candidate_anchors:
            self.logger.debug("Scheda SR strict check failed: missing Media/Min/Max in order")
            return None

        def score_anchor_strict(anchor_row: int) -> int:
            # Higher is better: sum numeric cells across the three rows
            rows_map = grouped[anchor_row]
            score = 0
            for label in ("Media", "Min", "Max"):
                score += sum(1 for v in rows_map[label].values() if v is not None)
            return score

        best_anchor = max(candidate_anchors, key=score_anchor_strict)
        best_rows = {k: v for k, v in grouped[best_anchor].items() if k in {"Media", "Min", "Max"}}
        best_headers = headers_by_anchor[best_anchor]

        # 4) If we only captured Min/Max but not Media, try to look in a tight window around them to add missing labels
        if "Media" not in best_rows and best_rows:
            row_indices = sorted(
                [ri for ri, lb in label_hits if lb in {"media", "min", "max"}]
            )
            # Pick a representative row from this anchor's captured rows
            representative_row = None
            for ri, lb in label_hits:
                if lb.capitalize() in best_rows:
                    representative_row = ri
                    break
            if representative_row is None:
                representative_row = list(best_rows.values())[0]
            # Scan a local window of ±5 rows to fill gaps
            window_lo = max(1, (representative_row if isinstance(representative_row, int) else 1) - 5)
            window_hi = min(max_rows, (representative_row if isinstance(representative_row, int) else 1) + 5)
            header_row_idx = best_anchor
            header_map = mapping_by_anchor.get(header_row_idx, {})
            for r in range(window_lo, window_hi + 1):
                row_iter = ws.iter_rows(min_row=r, max_row=r, max_col=max_cols)
                row_cells = next(row_iter, tuple())
                label = self._detect_summary_label(row_cells)
                if label in {"media", "min", "max"} and label.capitalize() not in best_rows:
                    # Ensure we still use the same header anchor
                    # (nearest above r must be best_anchor)
                    prior = [h for h in header_candidates if h[0] < r]
                    if prior and prior[-1][0] == best_anchor:
                        rv: Dict[str, Optional[float]] = {}
                        for h in best_headers:
                            col_idx = prior[-1][1][h]
                            rv[h] = self._parse_number(ws.cell(row=r, column=col_idx).value)
                        best_rows[label.capitalize()] = rv

        notes = self._collect_notes(ws, best_anchor)
        self.logger.debug(
            "Scheda SR summary extracted (rows: %s, headers: %s, notes=%d) using header row %d",
            list(best_rows.keys()),
            best_headers,
            len(notes),
            best_anchor,
        )
        return SchedaSummary(headers=best_headers, rows=best_rows, notes=notes)

    # ------------------------------------------------------------------
    def _extract_collaudo_media(self, wb) -> Optional[CollaudoSummary]:
        self.logger.debug("Attempting to extract Collaudo SR media row")

        # Be tolerant to sheet name variations: accept any sheet whose normalized
        # name contains 'collaudo' (handles trailing spaces, case differences, etc.)
        collaudo_sheet_name = None
        for s in wb.sheetnames:
            normalized = self._normalize_text(s)
            if "collaudo" in normalized or normalized == "collaudo":
                collaudo_sheet_name = s
                break

        if not collaudo_sheet_name:
            self.logger.debug("Workbook missing any sheet with name containing 'Collaudo'")
            return None

        ws: Worksheet = wb[collaudo_sheet_name]

        # NEW APPROACH: Find "media" in first column (column A), then locate headers above
        max_rows = min(ws.max_row or 0, 200)
        max_cols = min(ws.max_column or 0, 200)
        
        # Step 1: Find all rows with "media" in first column
        media_candidates: List[int] = []
        for row_idx in range(1, max_rows + 1):
            first_cell_value = ws.cell(row=row_idx, column=1).value
            normalized = self._normalize_text(first_cell_value)
            if normalized == "media" or "media" in normalized:
                media_candidates.append(row_idx)

        if not media_candidates:
            self.logger.debug("Media label not found in first column of Collaudo sheet")
            return None

        best_values: Optional[Dict[str, Optional[float]]] = None
        best_headers: Optional[List[str]] = None
        best_score = -1
        best_row = None

        # Step 2: For each media row candidate, find headers above and extract data
        for media_row in media_candidates:
            # Search for headers in rows above the media row (up to 20 rows above)
            header_search_start = max(1, media_row - 20)
            dynamic_column_map = self._resolve_collaudo_columns(ws, media_row, max_cols)
            
            values: Dict[str, Optional[float]] = {}
            headers: List[str] = []
            numeric_count = 0
            
            for header, default_col_idx in self._COLLAUDO_COLUMNS:
                headers.append(header)
                resolved_col_idx = dynamic_column_map.get(header, default_col_idx)
                if resolved_col_idx and resolved_col_idx <= (ws.max_column or 0):
                    cell_value = ws.cell(row=media_row, column=resolved_col_idx).value
                else:
                    if header not in dynamic_column_map:
                        self.logger.debug("Collaudo column '%s' not found; using default index %s", header, default_col_idx)
                    cell_value = None
                parsed = self._parse_number(cell_value)
                values[header] = parsed
                if parsed is not None:
                    numeric_count += 1

            if numeric_count > best_score:
                best_score = numeric_count
                best_values = values
                best_headers = headers
                best_row = media_row

        if not best_values or best_score <= 0:
            self.logger.debug("Collaudo SR media candidate rows contained no numeric data")
            return None

        self.logger.debug(
            "Collaudo SR media row selected (row %s) with %d numeric columns",
            best_row,
            best_score,
        )
        return CollaudoSummary(headers=best_headers or [], values=best_values)

    # ------------------------------------------------------------------
    def _resolve_collaudo_columns(self, ws: Worksheet, media_row_idx: int, max_cols: int) -> Dict[str, int]:
        column_map: Dict[str, int] = {}
        search_start = max(1, media_row_idx - 10)
        for row_idx in range(search_start, media_row_idx):
            row_iter = ws.iter_rows(min_row=row_idx, max_row=row_idx, max_col=max_cols)
            row_cells = next(row_iter, tuple())
            for cell in row_cells:
                normalized = self._normalize_text(getattr(cell, "value", None))
                if not normalized:
                    continue
                alias = self._COLLAUDO_HEADER_ALIASES.get(normalized)
                if not alias:
                    continue
                col_idx = self._column_index(cell)
                if col_idx is None:
                    continue
                column_map.setdefault(alias, col_idx)

        if not column_map:
            self.logger.debug("Collaudo header columns not detected dynamically; relying on defaults")
        else:
            self.logger.debug("Collaudo columns resolved dynamically: %s", column_map)
        return column_map

    # ------------------------------------------------------------------
    def _locate_scheda_headers(self, ws: Worksheet) -> tuple[int, Dict[str, int]]:
        # Backwards compatibility for callers; keep last header row
        candidates = self._find_all_scheda_header_rows(ws)
        if not candidates:
            return 0, {}
        return candidates[-1]

    def _find_all_scheda_header_rows(self, ws: Worksheet) -> List[tuple[int, Dict[str, int]]]:
        candidates: List[tuple[int, Dict[str, int]]] = []
        max_rows = min(ws.max_row, 200)
        max_cols = min(ws.max_column, 50)

        for row_idx in range(1, max_rows + 1):
            current_map: Dict[str, int] = {}
            row_iter = ws.iter_rows(min_row=row_idx, max_row=row_idx, max_col=max_cols)
            row_cells = next(row_iter, tuple())
            for cell in row_cells:
                normalized = self._normalize_text(cell.value)
                if not normalized:
                    continue
                alias = self._SCHEda_HEADER_ALIASES.get(normalized)
                if alias and alias not in current_map:
                    col_idx = self._column_index(cell)
                    if col_idx is not None:
                        current_map[alias] = col_idx

            if current_map:
                candidates.append((row_idx, current_map))

        return candidates

    # ------------------------------------------------------------------
    def _detect_summary_label(self, row_cells) -> str:
        if row_cells and isinstance(row_cells[0], tuple):
            # Flatten when iter_rows produced a tuple container inside a list
            iterable = (cell for group in row_cells for cell in group)
        else:
            iterable = iter(row_cells)

        for cell in iterable:
            normalized = self._normalize_text(getattr(cell, "value", None))
            if not normalized:
                continue
            # Accept common variations and substrings for labels
            if (normalized == "media" or normalized.startswith("med") or "media" in normalized):
                return "media"
            if (
                normalized == "min"
                or normalized.startswith("minim")
                or normalized in {"minimo", "minimi", "min."}
            ):
                return "min"
            if (
                normalized == "max"
                or normalized.startswith("massim")
                or normalized in {"massimo", "massimi", "max."}
            ):
                return "max"
        return ""

    # ------------------------------------------------------------------
    def _collect_notes(self, ws: Worksheet, header_row_idx: int) -> List[str]:
        notes: List[str] = []
        max_row = min(ws.max_row, header_row_idx + 10)
        for row_idx in range(header_row_idx + 1, max_row + 1):
            parts: List[str] = []
            # Notes are generally stored in columns B-E
            for col_idx in range(2, 6):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                text = str(value).strip()
                if text:
                    parts.append(text)
            if parts:
                note_line = " ".join(parts)
                if note_line not in notes:
                    notes.append(note_line)
        return notes

    # ------------------------------------------------------------------
    @staticmethod
    def _parse_number(value: object) -> Optional[float]:
        if value is None or value == "":
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            text = text.replace("\u00a0", "")
            # Handle European decimal separators (e.g. 1.234,56)
            if "," in text and "." in text:
                text = text.replace(".", "").replace(",", ".")
            elif "," in text:
                text = text.replace(",", ".")
            try:
                return float(text)
            except ValueError:
                return None
        return None

    # ------------------------------------------------------------------
    @staticmethod
    def _normalize_text(value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip().lower()
        text = text.replace("ø", "o").replace("φ", "o")
        text = re.sub(r"[^a-z0-9%]+", "", text)
        return text

    @staticmethod
    def _column_index(cell) -> Optional[int]:
        """Return a numeric column index for *cell*, handling merged cells safely."""
        col_idx = getattr(cell, "col_idx", None)
        if isinstance(col_idx, int):
            return col_idx

        column = getattr(cell, "column", None)
        if isinstance(column, int):
            return column
        if isinstance(column, str):
            try:
                return column_index_from_string(column)
            except ValueError:
                return None
        return None

    def get_raw_sheets_data(
        self, test_number: str
    ) -> Optional[Dict[str, Any]]:
        """Get all Scheda and Collaudo sheets from source workbook.
        
        Args:
            test_number: Test number (e.g., "27971A")
            
        Returns:
            Dict with 'scheda' and 'collaudo' lists of worksheets, or None if workbook not found
        """
        # Locate source workbook
        location_result = self.locate_workbook(test_number)
        if not location_result:
            logger.warning(f"Could not locate workbook for test {test_number}")
            return None
        
        source_path = location_result.path
        
        # Convert .xls to .xlsx if needed
        if source_path.suffix.lower() == ".xls":
            try:
                converted_path = self._convert_xls_to_xlsx(source_path)
                if not converted_path:
                    logger.error(f"Failed to convert .xls to .xlsx for test {test_number}")
                    return None
                source_path = converted_path
            except Exception as e:
                logger.error(f"Failed to convert .xls to .xlsx for test {test_number}: {e}")
                return None
        
        # Load source workbook - CRITICAL: data_only=False to preserve formulas!
        try:
            source_wb = openpyxl.load_workbook(source_path, data_only=False, keep_vba=False)
        except Exception as e:
            logger.error(f"Failed to load workbook {source_path}: {e}")
            return None
        
        # Find all Scheda and Collaudo sheets
        scheda_sheets = []
        collaudo_sheets = []
        
        for sheet_name in source_wb.sheetnames:
            normalized = sheet_name.lower().replace(" ", "").replace("_", "")
            if "scheda" in normalized:
                scheda_sheets.append(source_wb[sheet_name])
            elif "collaudo" in normalized:
                collaudo_sheets.append(source_wb[sheet_name])
        
        logger.info(f"Found {len(scheda_sheets)} Scheda and {len(collaudo_sheets)} Collaudo sheets for test {test_number}")
        
        return {
            'scheda': scheda_sheets,
            'collaudo': collaudo_sheets,
            'workbook': source_wb  # Keep reference to close later
        }
    
    @staticmethod
    def _copy_sheet_with_formatting(
        source_sheet: Worksheet, target_sheet: Worksheet
    ) -> None:
        """Copy worksheet with full formatting including formulas.
        
        Args:
            source_sheet: Source worksheet
            target_sheet: Target worksheet
        """
        # Copy dimensions
        target_sheet.sheet_format.defaultRowHeight = source_sheet.sheet_format.defaultRowHeight
        target_sheet.sheet_format.defaultColWidth = source_sheet.sheet_format.defaultColWidth
        
        # Copy column widths
        for col_letter, col_dim in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[col_letter].width = col_dim.width
        
        # Copy row heights
        for row_num, row_dim in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[row_num].height = row_dim.height
        
        # Copy cells with formatting
        for row in source_sheet.iter_rows():
            for cell in row:
                target_cell = target_sheet[cell.coordinate]
                
                # Copy value or formula
                if cell.data_type == 'f':  # Formula
                    target_cell.value = cell.value
                else:
                    target_cell.value = cell.value
                
                # Copy formatting
                if cell.has_style:
                    target_cell.font = copy(cell.font)
                    target_cell.border = copy(cell.border)
                    target_cell.fill = copy(cell.fill)
                    target_cell.number_format = cell.number_format
                    target_cell.protection = copy(cell.protection)
                    target_cell.alignment = copy(cell.alignment)
        
        # Copy merged cells
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))

    # ------------------------------------------------------------------
    def _extract_raw_sheets(self, workbook_path: Path) -> List[Dict[str, Any]]:
        """Extract raw data from all relevant sheets (Scheda/Collaudo/Carichi)."""
        self.logger.debug("Extracting raw sheets from %s", workbook_path)
        
        # Handle XLS conversion if needed
        workbook_to_read = workbook_path
        temp_converted_path = None
        
        if workbook_path.suffix.lower() == ".xls":
            temp_converted_path = self._convert_xls_to_xlsx(workbook_path)
            if temp_converted_path:
                workbook_to_read = temp_converted_path
            else:
                return []

        raw_sheets = []
        wb = None
        try:
            # Load with data_only=True to get values, but we might miss formulas.
            # For reporting, values are usually preferred.
            wb = openpyxl.load_workbook(workbook_to_read, data_only=True)
            
            for sheet_name in wb.sheetnames:
                normalized = self._normalize_text(sheet_name)
                if any(k in normalized for k in ["scheda", "collaudo", "carichi"]):
                    ws = wb[sheet_name]
                    sheet_data = {
                        "name": sheet_name,
                        "values": [],
                        "merges": [],
                        "col_widths": {},
                        "row_heights": {}
                    }
                    
                    # Extract values
                    for row in ws.iter_rows(values_only=True):
                        sheet_data["values"].append(list(row))
                        
                    # Extract merges
                    for merged_range in ws.merged_cells.ranges:
                        sheet_data["merges"].append(str(merged_range))
                        
                    # Extract column widths
                    for col_letter, col_dim in ws.column_dimensions.items():
                        sheet_data["col_widths"][col_letter] = col_dim.width
                        
                    # Extract row heights
                    for row_idx, row_dim in ws.row_dimensions.items():
                        sheet_data["row_heights"][row_idx] = row_dim.height
                        
                    raw_sheets.append(sheet_data)
                    self.logger.debug("Extracted raw sheet: %s", sheet_name)
                    
        except Exception as e:
            self.logger.error("Error extracting raw sheets from %s: %s", workbook_path, e)
        finally:
            if wb:
                wb.close()
            if temp_converted_path and temp_converted_path.exists():
                temp_converted_path.unlink(missing_ok=True)
                
        return raw_sheets

